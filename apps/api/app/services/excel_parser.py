"""
Excel parser — openpyxl + pandas 2.x
"""

import io
from typing import Any
import pandas as pd
import openpyxl


def parse_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """
    Parse an Excel or CSV file and extract all text/numeric content
    as a structured dict for LLM normalization.
    """
    result: dict[str, Any] = {
        "filename": filename,
        "sheets": [],
        "raw_text": "",
    }

    if filename.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=str)
        raw = df.to_string(index=False, header=False)
        result["sheets"].append({"name": "Sheet1", "data": df.values.tolist()})
        result["raw_text"] = raw
        return result

    # Excel
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    all_text_parts: list[str] = []

    for sheet_name in wb.sheetnames[:5]:  # max 5 sheets
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        text_parts: list[str] = []

        for row in ws.iter_rows(max_row=200, values_only=True):
            non_null = [str(c) for c in row if c is not None and str(c).strip()]
            if non_null:
                rows.append(list(row))
                text_parts.append(" | ".join(non_null))

        result["sheets"].append({
            "name": sheet_name,
            "data": rows,
        })
        all_text_parts.append(f"=== Feuille: {sheet_name} ===\n" + "\n".join(text_parts))

    result["raw_text"] = "\n\n".join(all_text_parts)
    return result
